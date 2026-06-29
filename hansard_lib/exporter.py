"""Builds the exportable .xlsx workbooks (and a .zip of per-PDF workbooks)."""
from __future__ import annotations

import io
import zipfile
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

TRANSCRIPT_COLS = [
    "Order", "Page", "Speaker (As Printed)", "Matched Name", "Constituency",
    "Jawatan", "Kementerian", "Speech Text",
]
COL_WIDTHS = {
    "Order": 7, "Page": 8, "Speaker (As Printed)": 32, "Matched Name": 32,
    "Constituency": 18, "Jawatan": 26, "Kementerian": 28, "Speech Text": 90,
    "PDF Source": 22, "Sitting Date": 14, "Cabinet Snapshot": 22,
}


def _style_sheet(ws, ncols: int, wrap_last: bool = True):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")
    for col_idx in range(1, ncols + 1):
        header = ws.cell(row=1, column=col_idx).value
        width = COL_WIDTHS.get(header, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    if wrap_last:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_metadata_sheet(writer, metadata_rows: list[dict], sheet_name: str = "Metadata"):
    meta_df = pd.DataFrame(metadata_rows)
    meta_df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    _style_sheet(ws, len(meta_df.columns), wrap_last=False)
    for col_idx in range(1, len(meta_df.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 26


def build_single_pdf_workbook(transcript_df: pd.DataFrame, metadata: dict) -> bytes:
    cols = [c for c in TRANSCRIPT_COLS if c in transcript_df.columns]
    out_df = transcript_df[cols]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        out_df.to_excel(writer, sheet_name="Transcript", index=False)
        ws = writer.sheets["Transcript"]
        _style_sheet(ws, len(cols))
        _write_metadata_sheet(writer, [metadata])
    return buf.getvalue()


def build_combined_workbook(
    transcript_df_all: pd.DataFrame, per_pdf_metadata: list[dict]
) -> bytes:
    cols_combined = [c for c in (["PDF Source", "Sitting Date"] + TRANSCRIPT_COLS) if c in transcript_df_all.columns]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        combined = transcript_df_all[cols_combined]
        combined.to_excel(writer, sheet_name="All Combined", index=False)
        _style_sheet(writer.sheets["All Combined"], len(cols_combined))

        for label in transcript_df_all["PDF Source"].unique():
            sub = transcript_df_all[transcript_df_all["PDF Source"] == label]
            cols = [c for c in TRANSCRIPT_COLS if c in sub.columns]
            sheet_name = _safe_sheet_name(label)
            sub[cols].to_excel(writer, sheet_name=sheet_name, index=False)
            _style_sheet(writer.sheets[sheet_name], len(cols))

        _write_metadata_sheet(writer, per_pdf_metadata)
    return buf.getvalue()


def build_zip_of_individual_workbooks(
    transcript_df_all: pd.DataFrame, per_pdf_metadata: list[dict]
) -> bytes:
    zip_buf = io.BytesIO()
    meta_by_label = {m["PDF Source"]: m for m in per_pdf_metadata}
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for label in transcript_df_all["PDF Source"].unique():
            sub = transcript_df_all[transcript_df_all["PDF Source"] == label]
            metadata = meta_by_label.get(label, {"PDF Source": label})
            wb_bytes = build_single_pdf_workbook(sub, metadata)
            fname = _safe_filename(label) + ".xlsx"
            zf.writestr(fname, wb_bytes)
    return zip_buf.getvalue()


def matching_xlsx_filename(pdf_label: str) -> str:
    """Public helper: the .xlsx filename that matches a given source PDF's
    filename, e.g. "DR-19122022.pdf" -> "DR-19122022.xlsx"."""
    return _safe_filename(pdf_label) + ".xlsx"


def _strip_pdf_ext(name: str) -> str:
    s = str(name)
    if s.lower().endswith(".pdf"):
        s = s[: -len(".pdf")]
    return s


def _safe_sheet_name(name: str) -> str:
    bad = set('[]:*?/\\')
    cleaned = "".join(c for c in _strip_pdf_ext(name) if c not in bad)
    return cleaned[:31] or "Sheet"


def _safe_filename(name: str) -> str:
    bad = set('[]:*?/\\<>|"')
    cleaned = "".join(c for c in _strip_pdf_ext(name) if c not in bad)
    return cleaned.replace(" ", "_") or f"hansard_{datetime.now():%Y%m%d%H%M%S}"
